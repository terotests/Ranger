#!/usr/bin/env python3
"""Generate Milestone-4 fixtures for the DataGrid / XLSX workbook viewer."""
from __future__ import annotations

import os
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle

ROOT = Path(__file__).resolve().parent.parent / "fixtures"


def recompress_stored(path: Path):
    """Ranger ZipReader expects STORED entries (same as make_fixtures.py)."""
    import zipfile, io, tempfile
    buf = io.BytesIO()
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_STORED) as zout:
        for info in zin.infolist():
            zout.writestr(info.filename, zin.read(info.filename))
    path.write_bytes(buf.getvalue())


def save_wb(wb: Workbook, path: Path):
    wb.save(path)
    recompress_stored(path)


def write_formats():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formats"
    ws["A1"] = "Kind"
    ws["B1"] = "Raw"
    ws["C1"] = "Display"
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    ws["C1"].font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E79")
    for col in "ABC":
        ws[f"{col}1"].fill = fill
    rows = [
        ("integer", 42, "0"),
        ("decimal", 1234.5, "0.00"),
        ("thousands", 1234567.89, "#,##0.00"),
        ("percent", 0.256, "0.00%"),
        ("currency_usd", 1999.5, "$#,##0.00"),
        ("currency_eur", 1999.5, "€#,##0.00"),
        ("scientific", 12345.0, "0.00E+00"),
        ("date_iso", date(2024, 3, 15), "YYYY-MM-DD"),
        ("date_eu", date(2024, 3, 15), "DD.MM.YYYY"),
        ("date_us", date(2024, 3, 15), "MM/DD/YYYY"),
        ("time", datetime(1899, 12, 31, 14, 30, 5), "HH:MM:SS"),
        ("negative", -42.5, "#,##0.00;-#,##0.00;-"),
        ("zero", 0, "#,##0.00;-#,##0.00;-"),
        ("text", "hello", "@"),
    ]
    for i, (kind, raw, fmt) in enumerate(rows, start=2):
        ws.cell(i, 1, kind)
        cell = ws.cell(i, 2, raw)
        cell.number_format = fmt
        ws.cell(i, 3, raw)
        ws.cell(i, 3).number_format = fmt
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.freeze_panes = "A2"
    save_wb(wb, ROOT / "formats.xlsx")
    print("wrote", ROOT / "formats.xlsx")


def write_merged():
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws.merge_cells("A1:D1")
    ws["A1"] = "Styled merged header"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="2E75B6")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A3:C3")
    ws["A3"] = "Horizontal merge"
    ws.merge_cells("A5:A7")
    ws["A5"] = "Vertical"
    ws.merge_cells("B5:D7")
    ws["B5"] = "2D merge block"
    ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B5"].fill = PatternFill("solid", fgColor="FFF2CC")
    save_wb(wb, ROOT / "merged.xlsx")
    print("wrote", ROOT / "merged.xlsx")


def write_business():
    wb = Workbook()
    # --- Sales ---
    ws = wb.active
    ws.title = "Sales"
    headers = ["Date", "Region", "Product", "Qty", "UnitPrice", "Revenue", "Margin"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    regions = ["North", "South", "East", "West"]
    products = ["Widget", "Gadget", "Doohickey", "Gizmo"]
    for r in range(2, 202):
        d = date(2024, 1 + ((r - 2) % 12), 1 + ((r - 2) % 27))
        region = regions[(r - 2) % 4]
        product = products[(r - 2) % 4]
        qty = 1 + ((r * 3) % 40)
        price = round(4.5 + ((r % 17) * 1.25), 2)
        revenue = round(qty * price, 2)
        margin = round(0.12 + ((r % 9) * 0.01), 2)
        ws.cell(r, 1, d).number_format = "YYYY-MM-DD"
        ws.cell(r, 2, region)
        ws.cell(r, 3, product)
        ws.cell(r, 4, qty).number_format = "0"
        ws.cell(r, 5, price).number_format = "$#,##0.00"
        ws.cell(r, 6, revenue).number_format = "$#,##0.00"
        ws.cell(r, 7, margin).number_format = "0.00%"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:G201"
    for i, w in enumerate([12, 10, 12, 8, 12, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Costs ---
    wc = wb.create_sheet("Costs")
    wc.merge_cells("A1:E1")
    wc["A1"] = "Operating Costs"
    wc["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    wc["A1"].fill = PatternFill("solid", fgColor="C65911")
    wc["A1"].alignment = Alignment(horizontal="center")
    for i, h in enumerate(["Category", "Jan", "Feb", "Mar", "_helper"], 1):
        cell = wc.cell(2, i, h)
        cell.font = Font(bold=True)
    cats = ["Rent", "Payroll", "Marketing", "Logistics", "R&D"]
    for i, cat in enumerate(cats, 3):
        wc.cell(i, 1, cat)
        for c in range(2, 5):
            wc.cell(i, c, 1000 + i * 120 + c * 35).number_format = "$#,##0.00"
        wc.cell(i, 5, f"=B{i}+C{i}+D{i}")  # helper formula
    wc.column_dimensions["E"].hidden = True

    # --- Forecast ---
    wf = wb.create_sheet("Forecast")
    wf["A1"] = "Metric"
    wf["B1"] = "Value"
    wf["A1"].font = Font(bold=True)
    wf["B1"].font = Font(bold=True)
    wf["A2"] = "BaseUnits"
    wf["B2"] = 1000
    wf["A3"] = "Growth"
    wf["B3"] = 0.08
    wf["B3"].number_format = "0.00%"
    wf["A4"] = "ForecastUnits"
    wf["B4"] = "=B2*(1+B3)"
    wf["A5"] = "AvgPrice"
    wf["B5"] = "=AVERAGE(Sales!E2:E50)"
    wf["B5"].number_format = "$#,##0.00"
    wf["A6"] = "ForecastRev"
    wf["B6"] = "=B4*B5"
    wf["B6"].number_format = "$#,##0.00"
    wf["A7"] = "AbsCheck"
    wf["B7"] = "=ABS(B3-0.1)"
    wf["A8"] = "Flag"
    wf["B8"] = '=IF(B3>0.05,"High","Low")'
    wf["A9"] = "SalesSum"
    wf["B9"] = "=SUM(Sales!F2:F100)"
    wf["B9"].number_format = "$#,##0.00"
    wf["A10"] = "CrossRef"
    wf["B10"] = "=Costs!B3"

    # --- Summary ---
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Business Workbook — Summary KPIs"
    ws2["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 32
    ws2["A3"] = "Total Revenue (sample)"
    ws2["B3"] = "=Forecast!B9"
    ws2["B3"].number_format = "$#,##0.00"
    ws2["A4"] = "Forecast Revenue"
    ws2["B4"] = "=Forecast!B6"
    ws2["B4"].number_format = "$#,##0.00"
    ws2["A5"] = "Growth"
    ws2["B5"] = "=Forecast!B3"
    ws2["B5"].number_format = "0.00%"
    ws2["A6"] = "Rent (Jan)"
    ws2["B6"] = "=Costs!B3"
    ws2["B6"].number_format = "$#,##0.00"
    for r in range(3, 7):
        ws2.cell(r, 3, r * 10)
    ws2.conditional_formatting.add(
        "C3:C6",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )
    # Hidden helper sheet
    wh = wb.create_sheet("Scratch")
    wh["A1"] = "hidden helper"
    wh.sheet_state = "hidden"

    save_wb(wb, ROOT / "business-workbook.xlsx")
    print("wrote", ROOT / "business-workbook.xlsx")


def write_sparse(n=100000):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse100k"
    ws["A1"] = "Row"
    ws["B1"] = "Marker"
    for i in range(1, n + 1):
        if i == 1 or i % 500 == 0 or i < 20:
            ws.cell(i, 1, i)
            ws.cell(i, 2, f"m-{i}")
    # Force dimension via last marker
    ws.cell(n, 1, n)
    ws.cell(n, 2, f"m-{n}")
    save_wb(wb, ROOT / "sparse100k.xlsx")
    print("wrote", ROOT / "sparse100k.xlsx")


def write_formula_fixture():
    wb = Workbook()
    # Input sheet
    inp = wb.active
    inp.title = "Input"
    inp["A1"] = 10
    inp["A2"] = 20
    inp["A3"] = 30

    # Calc sheet — relative, absolute, cross-sheet, IF
    ws = wb.create_sheet("Calc", 0)
    ws["A1"] = 10
    ws["A2"] = 20
    ws["B1"] = "=A1+A2"
    ws["C1"] = "=B1*2"
    ws["D1"] = "=SUM(A1:A2)"
    ws["E1"] = "=IF(A1>5,1,0)"
    ws["F1"] = "=A1/0"  # DIV/0
    ws["G1"] = "=COMPLEX(1,2)"  # unsupported → keep cached if present
    other = wb.create_sheet("Other")
    other["A1"] = 7
    ws["H1"] = "=Other!A1*3"
    ws["I1"] = "=$A$1+Input!A1"
    ws["J1"] = "=SUM(Input!A1:A3)"
    ws["K1"] = "=IF(J1>50,\"large\",\"small\")"

    # Errors sheet
    err = wb.create_sheet("Errors")
    err["A1"] = "=1/0"
    err["A2"] = "=IFERROR(A1,123)"

    # Quoted sheet name
    spaced = wb.create_sheet("My Sheet")
    spaced["A1"] = 5
    ws["L1"] = "='My Sheet'!A1*2"

    save_wb(wb, ROOT / "formulas.xlsx")
    print("wrote", ROOT / "formulas.xlsx")


def main():
    ROOT.mkdir(parents=True, exist_ok=True)
    write_formats()
    write_merged()
    write_business()
    write_formula_fixture()
    write_sparse(100000)
    # keep regenerating classic fixtures too
    os.system(f"python3 {Path(__file__).parent / 'make_fixtures.py'}")


if __name__ == "__main__":
    main()
