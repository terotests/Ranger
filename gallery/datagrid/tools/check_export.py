#!/usr/bin/env python3
"""Read the exported workbook with openpyxl — the independent half of the test.

`XlsxWriteTest` proves our writer and our loader agree with each other, which
is a weaker claim than it looks: two halves of one codebase can share a
misunderstanding. This reads the same file with a library that has never seen
our model, and fails loudly if anything it expects is missing.

    python3 gallery/datagrid/tools/check_export.py [path]

Skips (exit 0) when openpyxl is not installed, so the suite still runs on a
machine without it — but says so rather than pretending it checked.
"""
import os
import sys
import zipfile

DEFAULT = os.path.join(os.path.dirname(__file__), "..", "harness", "out",
                       "export", "roundtrip.xlsx")

failures = []
checks = 0


def check(name, got, want):
    global checks
    checks += 1
    if got == want:
        print(f"  PASS {name} (got {got!r})")
    else:
        print(f"  FAIL {name} (got {got!r} want {want!r})")
        failures.append(name)


def truthy(name, got):
    check(name, bool(got), True)


def main():
    path = os.path.abspath(sys.argv[1] if len(sys.argv) > 1 else DEFAULT)
    if not os.path.exists(path):
        print(f"no export at {path} — run npm run datagrid:export:test")
        return 1

    # A zip whose CRCs are wrong opens in some readers and not others, so it is
    # checked first and on its own: everything below would be meaningless.
    print("== the package is a valid zip ==")
    with zipfile.ZipFile(path) as z:
        bad = [i.filename for i in z.infolist() if _crc_bad(z, i)]
    check("every entry passes its CRC", bad, [])

    try:
        import openpyxl
    except ImportError:
        print("openpyxl is not installed — the independent read was SKIPPED")
        print(f"passed={checks - len(failures)} failed={len(failures)} (partial)")
        return 1 if failures else 0

    print("== openpyxl reads it ==")
    wb = openpyxl.load_workbook(path)
    check("both sheets", wb.sheetnames, ["Data", "Notes"])
    ws = wb["Data"]

    check("text", ws["A1"].value, "Region")
    check("escaped text", ws["A3"].value, "South & West")
    check("number stays a number", ws["B2"].value, 120)
    check("decimal", ws["C2"].value, 9.5)
    check("formula", ws["D2"].value, "=B2*C2")
    check("range formula", ws["D4"].value, "=SUM(D2:D3)")

    head = ws["A1"]
    truthy("bold", head.font.bold)
    check("font colour", head.font.color.rgb, "FF1F4E79")
    check("fill", head.fill.fgColor.rgb, "FFDDEBF7")
    check("horizontal alignment", head.alignment.horizontal, "center")
    truthy("wrap", head.alignment.wrap_text)
    check("border", ws["A1"].border.bottom.style, "thin")
    check("number format", ws["C2"].number_format, "$#,##0.00")
    check("text rotation", ws["A5"].alignment.textRotation, 45)

    check("column width", round(ws.column_dimensions["A"].width, 3), 19.375)
    truthy("hidden column", ws.column_dimensions["E"].hidden)
    check("row height", round(ws.row_dimensions[1].height, 1), 25.5)
    check("merge", [str(r) for r in ws.merged_cells.ranges], ["A7:C7"])
    check("freeze", ws.freeze_panes, "B2")

    check("external hyperlink", ws["A2"].hyperlink.target,
          "https://example.com/north")
    check("internal hyperlink", ws["A3"].hyperlink.location, "Notes!A1")

    dvs = sorted(ws.data_validations.dataValidation, key=lambda d: str(d.sqref))
    check("two validations", len(dvs), 2)
    if len(dvs) == 2:
        check("list rule", (dvs[0].type, str(dvs[0].sqref)), ("list", "A9:A12"))
        check("list values", dvs[0].formula1, '"north,south,east,west"')
        check("numeric rule", (dvs[1].type, dvs[1].operator), ("whole", "between"))
        check("its bounds", (dvs[1].formula1, dvs[1].formula2), ("1", "100"))
        check("its message", dvs[1].error, "Enter 1 to 100")

    print()
    print(f"passed={checks - len(failures)} failed={len(failures)}")
    if failures:
        print("FAILURES: " + ", ".join(failures))
        return 1
    print("ALL PASS")
    return 0


def _crc_bad(z, info):
    try:
        z.read(info.filename)
        return False
    except Exception:
        return True


if __name__ == "__main__":
    sys.exit(main())
