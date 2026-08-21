#!/usr/bin/env python3
"""Regenerate gallery/datagrid/fixtures/tables.xlsx.

A table (a ListObject) is a named rectangle whose columns have names, and it is
what lets a formula say what it means instead of where it is:

    =SUM(Table1[Amount])                   every amount
    =Table1[[#This Row],[CategoryID]]      the category on this row
    =Table1[@Amount]                       the same idea, written since 2010

Both spellings turn up in files people actually send. The long one is what
Excel 2007 wrote and what stays in a file written then; the short one is what
Excel writes today. Neither calculated here until the `xl/tables/*.xml` parts
were read, because nothing knew what `Table1` or `[Amount]` referred to.

The fixture carries two tables so the workbook-wide name lookup is exercised
rather than assumed, and the second one has a TOTALS ROW, which moves the last
data row up by one — a table where `[#Data]` and `[#All]` are different
rectangles and getting the two confused would show.

Every formula's expected value is written into the column beside it, so a test
can compare against the file rather than against its own arithmetic.

Run: python3 gallery/datagrid/tools/make_tables_fixture.py
"""
import os

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(
    os.path.abspath(__file__)))))
OUT = os.path.join(ROOT, "gallery", "datagrid", "fixtures", "tables.xlsx")

SALES = [
    ["ProductID", "CategoryID", "Amount", "Unit Price"],
    [1, 10, 100, 9.5],
    [2, 20, 250, 12.25],
    [3, 10, 75, 3.0],
    [4, 30, 410, 41.0],
]

# A second table, on its own sheet, with a totals row. `Staff[#Data]` is three
# rows; `Staff[#All]` is five.
STAFF = [
    ["Name", "Hours"],
    ["Ada", 38],
    ["Grace", 41],
    ["Alan", 35],
    ["Total", 114],
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    for row in SALES:
        ws.append(row)

    t1 = Table(displayName="Table1", ref="A1:D5")
    t1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(t1)

    # Formulas in F, what each one should come to in G. The 2007 spelling and
    # the 2010 one sit next to each other deliberately.
    checks = [
        ("=Table1[[#This Row],[CategoryID]]", 10),
        ("=Table1[@Amount]", 250),
        ("=SUM(Table1[Amount])", 835),
        ("=Table1[[#This Row],[Unit Price]]", 41.0),
        ("=COUNT(Table1[ProductID])", 4),
    ]
    ws["F1"] = "formula"
    ws["G1"] = "expected"
    for i, (formula, want) in enumerate(checks):
        ws.cell(row=i + 2, column=6, value=formula)
        ws.cell(row=i + 2, column=7, value=want)

    ws2 = wb.create_sheet("People")
    for row in STAFF:
        ws2.append(row)
    t2 = Table(displayName="Staff", ref="A1:B5")
    t2.totalsRowCount = 1
    t2.tableStyleInfo = TableStyleInfo(name="TableStyleLight1")
    ws2.add_table(t2)

    # Written on the OTHER sheet on purpose: a table name is workbook-wide, so
    # a formula on Sales must be able to reach a table on People.
    cross = [
        ("=SUM(Staff[Hours])", 114),
        ("=COUNT(Staff[Hours])", 3),
    ]
    for i, (formula, want) in enumerate(cross):
        ws.cell(row=i + 8, column=6, value=formula)
        ws.cell(row=i + 8, column=7, value=want)

    wb.save(OUT)
    print("wrote " + OUT)


if __name__ == "__main__":
    main()
