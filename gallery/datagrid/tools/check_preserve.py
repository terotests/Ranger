#!/usr/bin/env python3
"""Read the preserved workbooks with openpyxl — the independent half.

`XlsxPreserveTest` proves our writer keeps every part our own reader can see,
which is a weaker claim than it looks: two halves of one codebase can share a
misunderstanding about what a valid package is. A save-over merges the content
types, the package relationships, the workbook relationships and the
worksheet's own references, and getting any of those subtly wrong produces a
file that our loader is happy with and Excel refuses.

So this opens the same files with a library that has never seen our model, and
checks the contrast that the whole feature is about:

    writeOver  keeps the pictures, the comments, the theme and docProps
    write      keeps none of them, and is still a valid workbook

    python3 gallery/datagrid/tools/check_preserve.py

Skips (exit 0) when openpyxl is not installed, so the suite still runs on a
machine without it — but says so rather than pretending it checked.
"""
import os
import sys
import warnings
import zipfile

BIN = os.path.join(os.path.dirname(__file__), "..", "bin")

failures = []
checks = 0


def check(name, got, want):
    global checks
    checks += 1
    if got == want:
        print(f"  PASS  {name}")
    else:
        print(f"  FAIL  {name} (got {got!r} want {want!r})")
        failures.append(name)


def parts(path):
    with zipfile.ZipFile(path) as z:
        return set(z.namelist())


def main():
    try:
        import openpyxl
    except ImportError:
        print("check_preserve: openpyxl not installed — SKIPPED")
        return 0
    warnings.simplefilter("ignore")

    print("== the preserved files open in a reader that has never seen our model ==")
    preserved = sorted(
        f for f in os.listdir(BIN)
        if f.startswith("preserve-") and f.endswith(".xlsx")
    )
    if not preserved:
        print("  no preserved workbooks in gallery/datagrid/bin —"
              " run npm run datagrid:preserve:test first")
        return 1
    for name in preserved:
        path = os.path.join(BIN, name)
        try:
            wb = openpyxl.load_workbook(path)
            check(f"{name} opens", len(wb.sheetnames) > 0, True)
        except Exception as exc:  # noqa: BLE001 - report whatever it refuses on
            check(f"{name} opens", f"{type(exc).__name__}: {exc}", True)

    print("== and they still have what a plain save deletes ==")
    pics = os.path.join(BIN, "preserve-pics.xlsx")
    if os.path.exists(pics):
        wb = openpyxl.load_workbook(pics)
        drawn = sum(len(getattr(ws, "_images", [])) for ws in wb.worksheets)
        check("the pictures are on the sheet, not merely in the zip", drawn, 2)

    notes = os.path.join(BIN, "preserve-notes.xlsx")
    if os.path.exists(notes):
        check("the comments part survived",
              any(p.startswith("xl/comments") for p in parts(notes)), True)

    bw = os.path.join(BIN, "preserve-business-workbook.xlsx")
    if os.path.exists(bw):
        have = parts(bw)
        check("docProps survived", "docProps/core.xml" in have, True)
        check("the theme survived", "xl/theme/theme1.xml" in have, True)

    print("== and a clean save is genuinely clean, and genuinely valid ==")
    clean = os.path.join(BIN, "clean-images.xlsx")
    if os.path.exists(clean):
        have = parts(clean)
        check("no drawing", any(p.startswith("xl/drawings/") for p in have), False)
        check("no media", any(p.startswith("xl/media/") for p in have), False)
        try:
            wb = openpyxl.load_workbook(clean)
            check("but still a workbook", len(wb.sheetnames) > 0, True)
        except Exception as exc:  # noqa: BLE001
            check("but still a workbook", f"{type(exc).__name__}: {exc}", True)

    print()
    print(f"{checks} checks, {len(failures)} failed")
    if failures:
        for f in failures:
            print(f"  - {f}")
        return 1
    print("ALL PASS")
    return 0


if __name__ == "__main__":
    sys.exit(main())
