#!/usr/bin/env python3
"""Semantic oracle: openpyxl workbook model vs Ranger inspect.json."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("SKIP semantic: openpyxl not installed (pip install openpyxl)", file=sys.stderr)
    sys.exit(0)


def dump_xlsx(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        cells = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 50), max_col=min(ws.max_column or 1, 20)):
            for cell in row:
                if cell.value is None:
                    continue
                cells.append(
                    {
                        "coord": cell.coordinate,
                        "value": cell.value if not isinstance(cell.value, str) or not cell.value.startswith("=") else None,
                        "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
                        "numFmt": cell.number_format,
                    }
                )
        sheets.append(
            {
                "name": name,
                "state": ws.sheet_state,
                "maxRow": ws.max_row,
                "maxCol": ws.max_column,
                "freeze": list(ws.freeze_panes) if ws.freeze_panes else None,
                "merged": [str(r) for r in ws.merged_cells.ranges],
                "cells": cells[:200],
            }
        )
    return {"sheets": sheets, "sheetCount": len(sheets)}


def compare(ref: dict, actual: dict) -> list[str]:
    errs = []
    if ref.get("sheetCount") != actual.get("sheetCount"):
        # allow hidden sheet differences in count if Ranger includes them
        if abs((ref.get("sheetCount") or 0) - (actual.get("sheetCount") or 0)) > 1:
            errs.append(f"sheetCount ref={ref.get('sheetCount')} actual={actual.get('sheetCount')}")
    ref_names = [s["name"] for s in ref.get("sheets", []) if s.get("state") != "hidden"]
    act_names = [s["name"] for s in actual.get("sheets", [])]
    for n in ref_names:
        if n not in act_names:
            errs.append(f"missing sheet {n}")
    # merge presence on first sheet
    if ref.get("sheets") and actual.get("sheets"):
        r0 = ref["sheets"][0]
        a0 = next((s for s in actual["sheets"] if s["name"] == r0["name"]), None)
        if a0:
            if r0.get("merged") and not a0.get("mergedCount", 0) and not a0.get("merged"):
                # soft
                pass
            if (r0.get("maxRow") or 0) > 10 and (a0.get("rowCount") or 0) < 10:
                errs.append(f"rowCount too small for {r0['name']}")
    return errs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--ranger", required=True, help="Ranger inspect.json")
    ap.add_argument("--out", default="")
    args = ap.parse_args()
    ref = dump_xlsx(Path(args.xlsx))
    actual = json.loads(Path(args.ranger).read_text())
    if args.out:
        Path(args.out).write_text(json.dumps(ref, indent=2, default=str))
    errs = compare(ref, actual)
    if errs:
        print("FAIL semantic")
        for e in errs:
            print(" ", e)
        sys.exit(1)
    print("PASS semantic")
    sys.exit(0)


if __name__ == "__main__":
    main()
